"""
Борлуулалтын Excel файлыг импортлох команд.

Ашиглах:
    python manage.py import_sales_excel Бараа.xlsx
    python manage.py import_sales_excel Бараа.xlsx --dry-run        # Хадгалалгүй шалгах
    python manage.py import_sales_excel Бараа.xlsx --skip-existing  # Давхардсаныг алгасах

Логик:
  - Нэг огноо + нэг харилцагч = нэг Sale (SaleItem-ууд дотор нь)
  - Sale.status = DRAFT (банкны баримттай холбоогүй)
  - Ажилтан гараар BankTransaction.income_sale-г тохируулна
  - Нэг борлуулалт олон BankTransaction-тай холбогдож болно (касс + харилцах)
"""

import hashlib
import os
import sys
from collections import defaultdict
from datetime import datetime, date
from decimal import Decimal

# Django environment тохируулах (шууд python-оор ажиллуулахад)
if not os.environ.get('DJANGO_SETTINGS_MODULE'):
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    _project_root = os.path.dirname(os.path.dirname(os.path.dirname(_script_dir)))
    if _project_root not in sys.path:
        sys.path.insert(0, _project_root)
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'gotopa_project.settings')
    import django
    django.setup()

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction as db_transaction


# Excel баганы индексүүд
COL_NUM        = 0   # №
COL_DATE       = 1   # Огноо
COL_PRODUCT    = 2   # Барааны нэр
COL_UNIT       = 3   # х.нэгж
COL_QTY        = 4   # тоо хэмжээ
COL_UNIT_PRICE = 5   # нэгжийн үнэ
COL_TOTAL      = 6   # нийт үнэ
COL_CUSTOMER   = 7   # Харилцагчийн мэдээлэл
COL_SALESPERSON= 8   # Хүлээн авсан хүн
COL_PAY_TYPE   = 9   # Дансны төрөл (Касс / Харилцах)
COL_PAY_DONE   = 10  # Төлбөр дансанд орсон эсэх
COL_BANK       = 11  # Банк
COL_NOTE       = 12  # Тайлбар

# Монгол хэмжих нэгжийг Product.unit сонголтуудтай харгалзуулах
UNIT_MAP = {
    'ш':      'PIECE',
    'ширхэг': 'PIECE',
    'шир':    'PIECE',
    'хайрцаг':'BOX',
    'хайрцag':'BOX',
    'боодол': 'PACK',
    'багц':   'PACK',
    'кг':     'KG',
    'литр':   'LITER',
    'л':      'LITER',
    'метр':   'METER',
    'м':      'METER',
    'иж':     'SET',
    'set':    'SET',
}


def _unit(raw):
    """Монгол нэгжийг Product.unit код болгох."""
    if not raw:
        return 'PIECE'
    return UNIT_MAP.get(str(raw).strip().lower(), 'PIECE')


def _decimal(val):
    """Утгыг Decimal болгох (None → 0)."""
    if val is None:
        return Decimal('0')
    try:
        return Decimal(str(val))
    except Exception:
        return Decimal('0')


def _to_date(val):
    """datetime эсвэл date утгыг date болгох."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # str байвал parse хийх
    # АНХААР: '%m/%d/%Y' нь '%d/%m/%Y'-с өмнө байх ёстой —
    # учир нь "1/3/2026" гэх мэт тоонд хоёр формат хоёулаа таарна.
    # Хэрэв '%d/%m/%Y' эхэлж байвал 1/3/2026 → 3-р сарын 1 болж буруу уншина.
    # '%m/%d/%Y' эхэлж байвал 1/3/2026 → 1-р сарын 3 зөв уншина.
    for fmt in (
        '%Y.%m.%d',   # 2025.01.15
        '%d.%m.%Y',   # 15.01.2025
        '%Y-%m-%d',   # 2025-01-15
        '%Y/%m/%d',   # 2025/01/15
        '%m/%d/%Y',   # 1/3/2026 → 1-р сарын 3  ← энд байх ёстой
        '%d/%m/%Y',   # 15/01/2025
        '%d-%m-%Y',   # 15-01-2025
    ):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            pass
    return None


def _import_ref(sale_date, customer_name, product_names):
    """Давхардлаас сэргийлэх гарын үсэг үүсгэх."""
    raw = f"{sale_date}|{customer_name}|{'_'.join(sorted(product_names))}"
    return 'EXCEL-БАРАА-' + hashlib.md5(raw.encode('utf-8')).hexdigest()[:12]


class Command(BaseCommand):
    help = 'Борлуулалтын Excel файлыг Sale/SaleItem загварт импортлох'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Excel файлын зам (жишээ: Бараа.xlsx)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='Хадгалалгүйгээр зөвхөн шалгах',
        )
        parser.add_argument(
            '--skip-existing',
            action='store_true',
            default=False,
            help='Аль хэдийн импортолсон борлуулалтыг алгасах (default: алгасана)',
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='Хуудасны нэр (default: эхний хуудас)',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl суугаагүй байна: pip install openpyxl')

        # Windows консолын encoding-ийг UTF-8 болгох
        import sys
        if hasattr(sys.stdout, 'reconfigure'):
            try:
                sys.stdout.reconfigure(encoding='utf-8')
            except Exception:
                pass

        from main.models import Sale, SaleItem, Product, ProductCategory, Counterparty

        file_path = options['excel_file']
        dry_run   = options['dry_run']

        self.stdout.write(f"\n{'='*60}")
        self.stdout.write(f"[FILE] {file_path}")
        if dry_run:
            self.stdout.write("[DRY-RUN] Хадгалалгүй шалгах горим")
        self.stdout.write('='*60)

        # Excel нээх
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'Файл олдсонгүй: {file_path}')

        sheet_name = options['sheet'] or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise CommandError(f'Хуудас "{sheet_name}" олдсонгүй. Байгаа хуудаснууд: {wb.sheetnames}')

        ws = wb[sheet_name]
        self.stdout.write(f"[SHEET] {sheet_name}  |  {ws.max_row} мөр, {ws.max_column} багана\n")

        # ── 1. Мөрүүдийг уншиж, бүлэглэх ──────────────────────────────
        # Бүлгийн түлхүүр: (sale_date, customer_name)
        # Нэг бүлэг = нэг Sale
        groups = defaultdict(list)   # key → list of row-dicts
        skipped_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Хоосон мөр алгасах
            if not row[COL_DATE] and not row[COL_PRODUCT]:
                continue

            sale_date    = _to_date(row[COL_DATE])
            product_name = str(row[COL_PRODUCT] or '').strip()
            customer_raw = str(row[COL_CUSTOMER] or '').strip()
            salesperson  = str(row[COL_SALESPERSON] or '').strip()
            pay_type     = str(row[COL_PAY_TYPE] or '').strip()   # Касс / Харилцах
            pay_done     = str(row[COL_PAY_DONE] or '').strip()   # Тийм / None
            note         = str(row[COL_NOTE] or '').strip()
            unit_raw     = row[COL_UNIT]
            qty          = int(row[COL_QTY] or 1)
            unit_price   = _decimal(row[COL_UNIT_PRICE])
            total_price  = _decimal(row[COL_TOTAL]) or (unit_price * qty)

            if not sale_date or not product_name:
                self.stdout.write(self.style.WARNING(
                    f"  [WARN] Mor {row_idx}: ognoo esvel baraaany ner hooson -- algasav"
                ))
                skipped_rows += 1
                continue

            # Харилцагч хоосон бол "Бэлэн худалдан авагч" болгох
            customer_name = customer_raw if customer_raw else 'Бэлэн'

            key = (sale_date, customer_name)
            groups[key].append({
                'product_name': product_name,
                'unit':         _unit(unit_raw),
                'unit_raw':     str(unit_raw or 'ш'),
                'qty':          qty,
                'unit_price':   unit_price,
                'total_price':  total_price,
                'salesperson':  salesperson,
                'pay_type':     pay_type,
                'pay_done':     pay_done,
                'note':         note,
                'row_idx':      row_idx,
            })

        self.stdout.write(f"  Нийт бүлэг (борлуулалт): {len(groups)}")
        self.stdout.write(f"  Алгасагдсан мөр: {skipped_rows}\n")

        # ── 2. Бүлэг бүрийг Sale болгон оруулах ───────────────────────
        created_count  = 0
        skipped_count  = 0
        product_created = 0

        with db_transaction.atomic():
            for (sale_date, customer_name), rows in sorted(groups.items()):

                product_names = [r['product_name'] for r in rows]
                ref = _import_ref(sale_date, customer_name, product_names)

                # Давхардал шалгах
                if Sale.objects.filter(import_ref=ref).exists():
                    self.stdout.write(self.style.WARNING(
                        f"  [SKIP] {sale_date} / {customer_name} - аль хэдийн импортолсон, алгасав"
                    ))
                    skipped_count += 1
                    continue

                # Төлбөрийн хэлбэрүүдийг цуглуулах (Касс, Харилцах, Хоёулаа)
                pay_types = list(dict.fromkeys(
                    r['pay_type'] for r in rows if r['pay_type']
                ))
                pay_method_hint = ' + '.join(pay_types) if pay_types else ''

                # Борлуулагч нэр(үүд)
                salespersons = list(dict.fromkeys(
                    r['salesperson'] for r in rows if r['salesperson']
                ))
                salesperson_name = ', '.join(salespersons)

                # Тэмдэглэл цуглуулах
                notes_parts = list(dict.fromkeys(
                    r['note'] for r in rows if r['note']
                ))
                notes_text = '\n'.join(notes_parts)

                # Нийт дүн
                total = sum(r['total_price'] for r in rows)

                if not dry_run:
                    # Харилцагч get_or_create
                    if customer_name != 'Бэлэн':
                        counterparty, _ = Counterparty.objects.get_or_create(
                            name=customer_name,
                            defaults={
                                'counterparty_type': 'CUSTOMER',
                                'is_active': True,
                            }
                        )
                    else:
                        counterparty = None

                    # Sale үүсгэх
                    sale = Sale(
                        customer             = counterparty,
                        sale_date            = sale_date,
                        status               = 'DRAFT',
                        total_amount         = total,
                        paid_amount          = Decimal('0'),
                        salesperson_name     = salesperson_name,
                        expected_payment_method = pay_method_hint,
                        notes                = notes_text,
                        import_ref           = ref,
                    )
                    sale.save()

                    # SaleItem-үүд үүсгэх
                    for r in rows:
                        # Бүтээгдэхүүн хайх (нэрээр)
                        # Product.name unique биш тул filter().first() ашиглах
                        product = Product.objects.filter(
                            name=r['product_name'], is_active=True
                        ).first()
                        if product is None:
                            product = Product.objects.create(
                                name          = r['product_name'],
                                code          = _generate_product_code(r['product_name']),
                                unit          = r['unit'],
                                purchase_price= Decimal('0'),
                                selling_price = r['unit_price'],
                                is_active     = True,
                            )
                            product_created += 1

                        SaleItem.objects.create(
                            sale       = sale,
                            product    = product,
                            quantity   = r['qty'],
                            unit_price = r['unit_price'],
                            total_price= r['total_price'],
                        )

                    created_count += 1
                    self.stdout.write(self.style.SUCCESS(
                        f"  [OK] {sale.sale_number}  {sale_date}  "
                        f"{customer_name:<20}  {len(rows)} бараа  "
                        f"{total:>12,.0f}T  [{pay_method_hint}]"
                    ))
                else:
                    # Dry-run — зөвхөн мэдэгдэх
                    created_count += 1
                    self.stdout.write(
                        f"  [DRY] {sale_date}  {customer_name:<20}  "
                        f"{len(rows)} бараа  {total:>12,.0f}T  [{pay_method_hint}]"
                    )
                    for r in rows:
                        self.stdout.write(
                            f"         {r['product_name']}  "
                            f"x{r['qty']}  {r['unit_price']:,.0f}T"
                        )

            if dry_run:
                db_transaction.set_rollback(True)

        # ── 3. Дүгнэлт ────────────────────────────────────────────────
        self.stdout.write(f"\n{'='*60}")
        if dry_run:
            self.stdout.write(self.style.WARNING("[DRY-RUN] Өгөгдөл хадгалагдаагүй"))
        else:
            self.stdout.write(self.style.SUCCESS(
                f"[DONE]\n"
                f"   Борлуулалт үүссэн : {created_count}\n"
                f"   Алгасагдсан        : {skipped_count}\n"
                f"   Шинэ бүтээгдэхүүн : {product_created}"
            ))
        self.stdout.write('')
        self.stdout.write(
            "[INFO] Банкны баримттай холбох:\n"
            "   Dashboard -> Банкны гүйлгээ -> Гүйлгээ сонгоод -> "
            "'Борлуулалт' талбарт холбогдох борлуулалтыг зааж огно\n"
            "   Нэг борлуулалт олон гүйлгээтэй холбогдож болно (касс + харилцах)."
        )
        self.stdout.write('='*60 + '\n')


def _generate_product_code(name: str) -> str:
    """
    Бүтээгдэхүүний нэрнээс давтагдахгүй барааны код үүсгэх.
    Жишээ: 'Хонх жижиг' → 'PRD-ХОНХ-001'
    """
    from main.models import Product

    # Эхний 6 тэмдэгтийг авах (зай хасаж)
    prefix = name.strip()[:6].upper().replace(' ', '')
    base   = f'PRD-{prefix}'
    code   = base

    counter = 1
    while Product.objects.filter(code=code).exists():
        code = f'{base}-{counter:03d}'
        counter += 1

    return code


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Ашиглах: python main/management/commands/import_sales_excel.py <Бараа.xlsx> [--dry-run] [--skip-existing]')
        sys.exit(1)
    # management command-ийг дуурайж ажиллуулах
    from django.core.management import call_command
    kwargs = {}
    file_arg = sys.argv[1]
    if '--dry-run' in sys.argv:
        kwargs['dry_run'] = True
    if '--skip-existing' in sys.argv:
        kwargs['skip_existing'] = True
    call_command('import_sales_excel', file_arg, **kwargs)
